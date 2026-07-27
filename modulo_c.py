"""
modulo_c.py — Módulo C: Mapeo, cálculo y generación de output
Responsable: Lucía

Input:  list[dict] — salida de modulo_b.construir_entrada_modulo_c()
        26 columnas base + columnas auxiliares (Ancho CSV, Alto CSV, Alto final tapeta,
        Acabado del mueble abierto, Summary). Valores en lenguaje UI (CLAUDE.md §9).

Output principal:
    calcular_opciones(entrada) → list[dict]   para paso_2 de Módulo B
    generar_json_pedido(resultado) → list[dict]
    generar_excel_revision(resultado) → bytes (.xlsx)

Fuentes de verdad (no hardcodear valores aquí):
    data/catalogo.json          — códigos → dimensiones y designación ES
    data/mapeos_SKP_UI_SG.yaml  — conversión UI → código SG
    data/opciones_mueble.yaml   — qué opciones aplican a cada mueble
    data/reglas.yaml            — reglas de negocio entre opciones

"""

from __future__ import annotations

import functools
import json
from datetime import datetime
from io import BytesIO
from pathlib import Path

import yaml


# =============================================================================
# Rutas
# =============================================================================
_DATA_DIR        = Path(__file__).parent / "data"
_CATALOGO_PATH   = _DATA_DIR / "catalogo.json"
_MAPEOS_PATH     = _DATA_DIR / "mapeos_SKP_UI_SG.yaml"
_OPCIONES_PATH   = _DATA_DIR / "opciones_mueble.yaml"
_REGLAS_PATH     = _DATA_DIR / "reglas.yaml"
_AVISOS_PATH     = _DATA_DIR / "avisos.yaml"
_SCHEMA_PATH     = _DATA_DIR / "p_item_schema.yaml"


# =============================================================================
# Carga de datos (cached por proceso — sin dependencia de Streamlit)
# =============================================================================
@functools.lru_cache(maxsize=1)
def _cargar_datos() -> tuple[dict, dict, dict, dict]:
    """Carga los cuatro archivos de datos principales. Se cachea tras la primera llamada."""
    with _CATALOGO_PATH.open(encoding="utf-8") as f:
        catalogo = json.load(f)
    with _MAPEOS_PATH.open(encoding="utf-8") as f:
        mapeos = yaml.safe_load(f) or {}
    with _OPCIONES_PATH.open(encoding="utf-8") as f:
        op_mueble = yaml.safe_load(f) or {}
    with _REGLAS_PATH.open(encoding="utf-8") as f:
        reglas = yaml.safe_load(f) or {}
    return catalogo, mapeos, op_mueble, reglas


@functools.lru_cache(maxsize=1)
def _cargar_avisos() -> dict[str, str]:
    """Carga data/avisos.yaml sección modulo_c. Se cachea tras la primera llamada."""
    if not _AVISOS_PATH.exists():
        return {}
    with _AVISOS_PATH.open(encoding="utf-8") as f:
        return (yaml.safe_load(f) or {}).get("modulo_c") or {}


@functools.lru_cache(maxsize=1)
def _cargar_p_item_schema() -> dict:
    """Carga data/p_item_schema.yaml. Se cachea tras la primera llamada."""
    if not _SCHEMA_PATH.exists():
        return {}
    with _SCHEMA_PATH.open(encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def _p_item_defaults() -> dict:
    """Devuelve los valores estáticos del p_item definidos en p_item_schema.yaml.

    Filtra los marcadores __computed__ (campos calculados en tiempo de ejecución)
    y resuelve __today__ a la fecha actual en formato YYYY-MM-DD.
    """
    raw = (_cargar_p_item_schema().get("p_item_defaults") or {})
    today = datetime.today().strftime("%Y-%m-%d")
    return {
        k: (today if v == "__today__" else v)
        for k, v in raw.items()
        if v != "__computed__"
    }


def _p_variant_option_defaults() -> dict:
    """Devuelve los valores estáticos de p_variant_option definidos en p_item_schema.yaml."""
    raw = (_cargar_p_item_schema().get("p_variant_option_defaults") or {})
    return {k: v for k, v in raw.items() if v != "__computed__"}


# =============================================================================
# Constantes
# =============================================================================

# Tiradores por tipo (para op_300, op_301 y lógica de color)
_TIRADORES_INTEGRADOS = frozenset({"Round", "Square"})
_TIRADORES_SUPERFICIE = frozenset({"Curve", "Line", "Plantea"})
_TIRADORES_SIN_COLOR  = frozenset({"Touch Latch", "Prise de main"})

# (No hay constantes hardcodeadas aquí — todos los valores vienen de los YAMLs)
# op_700 forzadas/no_aplica → data/opciones_mueble.yaml (op_700.forzadas / op_700.no_aplica)
# Frases de avisos: strings directos en español — no se usan códigos cortos (AV0x)


# =============================================================================
# Helpers internos
# =============================================================================

def _build_indices(mapeos: dict) -> dict[str, dict[str, str]]:
    """
    Construye índices {ui → sg} para cada tabla de del_csv.
    Añade también {ui → skp} para op_300 (necesario para reglas de op_121).
    """
    del_csv = mapeos.get("del_csv") or {}
    result: dict[str, dict[str, str]] = {}
    for key, entries in del_csv.items():
        if not isinstance(entries, list):
            continue
        result[key] = {
            e["ui"]: e["sg"]
            for e in entries
            if "ui" in e and "sg" in e
        }
        if key == "op_300":
            result["op_300_ui_to_skp"] = {
                e["ui"]: str(e["skp"])
                for e in entries
                if "ui" in e and "skp" in e
            }
    return result


def _opt(numero: str | int, articulo: str, param1: str | int = "0") -> dict:
    """Construye un dict de opción para p_variant_options.
    Los valores reservados (p_param2, p_quantity) vienen de data/p_item_schema.yaml.
    """
    _vd = _p_variant_option_defaults()
    return {
        "p_option":   str(numero),
        "p_article":  str(articulo),
        "p_param1":   str(param1),
        "p_param2":   _vd.get("p_param2", "0"),
        "p_quantity": _vd.get("p_quantity", "1"),
    }


def _es_true(value) -> bool:
    return str(value).strip() == "True"


def _p_hinge_cat(code: str, op_mueble: dict) -> str:
    """Categoría p_hinge del mueble según opciones_mueble.yaml."""
    ph = op_mueble.get("p_hinge") or {}
    for cat in ("lee_csv", "nulo", "coulissant", "lift"):
        if code in (ph.get(cat) or []):
            return cat
    return ""


def _es_batiente(code: str, op_mueble: dict) -> bool:
    """True si el mueble tiene puertas batientes (→ sufijo H en Curve/Line)."""
    return _p_hinge_cat(code, op_mueble) in ("lee_csv", "nulo")


def _es_suspendido(code: str, op_mueble: dict) -> bool:
    """True si es mueble de pared (→ p_fastening = 'S')."""
    excepciones = (op_mueble.get("op_402") or {}).get("excepciones") or []
    return code in excepciones


# =============================================================================
# Cálculo de opciones por mueble
# =============================================================================

def _calcular_opciones_mueble(
    fila: dict,
    mapeos: dict,
    op_mueble: dict,
    reglas: dict,
    indices: dict,
) -> tuple[list[dict], list[dict], dict[str, str], list[str]]:
    """
    Calcula todas las opciones SG para un mueble.

    Devuelve:
        opciones_sg      — lista de dicts para p_variant_options (JSON)
        opc_adicionales  — lista de {etiqueta, valor, origen} para paso_2
        codigos          — dict {op_id → código_sg} para paso_2 y Excel
        avisos           — lista de códigos AV0x
    """
    code       = (fila.get("Código mueble") or "").strip()
    tirador_ui = (fila.get("Tirador")        or "").strip()
    gama_ui    = (fila.get("Gama del frente") or "").strip()

    opciones_sg: list[dict]     = []
    opc_adicionales: list[dict] = []
    codigos: dict[str, str]     = {}
    avisos: list[str]           = []

    forzadas_yaml = mapeos.get("forzadas") or {}
    reglas_c      = (reglas.get("modulo_c") or {})
    reglas_b      = (reglas.get("modulo_b") or {})

    def _sg(
        op_id: str,
        codigo: str,
        param1: str | int = "0",
        origen: str = "automatico",
        etiqueta: str = "",
    ) -> None:
        """Registra una opción en las tres estructuras de salida."""
        num = op_id.replace("op_", "")
        opciones_sg.append(_opt(num, codigo, param1))
        codigos[op_id] = codigo
        # param1 extra para op_231 (ancho reducido en mm)
        if op_id == "op_231" and str(param1) != "0":
            codigos["op_231_param1"] = str(param1)
        if etiqueta:
            opc_adicionales.append({
                "etiqueta": etiqueta,
                "valor": codigo,
                "origen": origen,
            })

    # ── Muebles abiertos (EOV/EOAVV): op_410/op_420 en lugar de frente/tirador ──
    codigos_mueble_abierto: set[str] = set(
        ((op_mueble.get("op_410") or {}).get("codigos")) or []
    )
    es_abierto = code in codigos_mueble_abierto

    # ── Tapetas (FF*/FFAL*): op_101 viene de "Acabado" en lugar de "Acabado del frente" ──
    codigos_tapeta: set[str] = set((op_mueble.get("tapetas") or {}).get("codigos") or [])

    # ── Rodapiés (SOC*): op_401 en lugar de op_100+op_101, viene de "Acabado" ──
    _rodapies_cfg = op_mueble.get("rodapiés") or {}
    codigos_rodapie: set[str] = set(_rodapies_cfg.get("codigos") or [])
    codigos_rodapie_sg: set[str] = set(_rodapies_cfg.get("codigos_sg") or [])
    es_rodapie = code in codigos_rodapie

    # ── Joues (J19*): op_621 (coloris panneaux) desde "Acabado". Sin op_100 ni op_101. ──
    codigos_joue: set[str] = set((op_mueble.get("joues") or {}).get("codigos") or [])
    es_joue = code in codigos_joue

    # ── Encimeras panel (PAN20LAMC, PAN20LINOC): op_621 desde "Acabado". ──
    codigos_enc_panel: set[str] = set((op_mueble.get("encimeras_panel") or {}).get("codigos") or [])
    es_enc_panel = code in codigos_enc_panel

    # ── Encimeras PDT (PDT20C, NRWC): op_601 + op_603 desde "Acabado". PDT20C: op_637=CAP. ──
    _enc_pdt_cfg: dict = op_mueble.get("encimeras_pdt") or {}
    codigos_enc_pdt: set[str] = set(_enc_pdt_cfg.get("codigos") or [])
    es_enc_pdt = code in codigos_enc_pdt

    es_encimera = es_enc_panel or es_enc_pdt

    if es_abierto:
        # op_410 — Acabado del mueble abierto (viene de la columna "Acabado")
        # op_420 — Acabado de la trasera (mismo código SG, forzada e invisible)
        color_abierto = (fila.get("Acabado") or "").strip()
        sg_410 = indices.get("op_410", {}).get(color_abierto, "")
        if sg_410:
            _sg("op_410", sg_410)  # visible en Configuración, no en Opciones adicionales
            _sg("op_420", sg_410)  # forzada e invisible, mismo código
    elif es_rodapie:
        # ── op_401 — Acabado del rodapié (mismo índice de valores que op_101) ────
        # No se envía op_100 (gama), SG lo determina a partir del código de artículo.
        acabado_rod = (fila.get("Acabado") or "").strip()
        sg_401 = indices.get("op_101", {}).get(acabado_rod, "")
        if sg_401:
            _sg("op_401", sg_401)
    elif es_joue:
        # ── op_621 — Coloris panneaux (mismo mapeo que op_101, sin op_100) ────
        acabado_joue = (fila.get("Acabado") or "").strip()
        sg_621 = indices.get("op_101", {}).get(acabado_joue, "")
        if sg_621:
            _sg("op_621", sg_621)
    elif es_enc_panel:
        # ── op_621 — Coloris panneaux (PAN20LAMC, PAN20LINOC). Mismo mapeo que op_101. ──
        acabado_enc = (fila.get("Acabado") or "").strip()
        sg_621_enc  = indices.get("op_101", {}).get(acabado_enc, "")
        if sg_621_enc:
            _sg("op_621", sg_621_enc)
    elif es_enc_pdt:
        # ── op_601 + op_603 — Coloris PdT y chant (PDT20C, NRWC). ──
        acabado_enc = (fila.get("Acabado") or "").strip()
        sg_601 = indices.get("op_601", {}).get(acabado_enc, "")
        if sg_601:
            _sg("op_601", sg_601)
        sg_603 = indices.get("op_603", {}).get(acabado_enc, "")
        if sg_603:
            _sg("op_603", sg_603)
        # ── op_637 — Canto encimera: forzada CAP solo en PDT20C. ──
        if code in set(_enc_pdt_cfg.get("op_637_forzada") or []):
            _sg("op_637", "CAP")
    else:
        # ── op_100 — Gama del frente (todos: O) ──────────────────────────────────
        sg_100 = indices.get("op_100", {}).get(gama_ui, "")
        if sg_100:
            _sg("op_100", sg_100)

        # ── op_101 — Acabado del frente ──────────────────────────────────────────
        # Tapetas: op_101 viene de "Acabado" (CSV directo). Resto: "Acabado del frente".
        acabado_col = "Acabado" if code in codigos_tapeta else "Acabado del frente"
        acabado = (fila.get(acabado_col) or "").strip()
        sg_101  = indices.get("op_101", {}).get(acabado, "")
        if sg_101:
            _sg("op_101", sg_101)

    # ── op_103, op_104 — EXCLUIDAS: SG las calcula en AS400 ─────────────────

    if not es_abierto:
        # ── op_200 — Color interior (todos: O excepto frentes sin mueble) ────────
        excl_200 = (op_mueble.get("op_200") or {}).get("excepciones") or []
        sg_200 = ""
        if code not in excl_200:
            interior = (fila.get("Color interior") or "").strip()
            sg_200   = indices.get("op_200", {}).get(interior, "")
            if sg_200:
                _sg("op_200", sg_200)

        # ── op_203, op_204 — Tapetas (FF*/FFAL*) únicamente ──────────────────────
        # Confirmado por Schmidt Groupe en validación VAC Recoletos (jun 2026): para
        # tapetas, CUBRO SÍ envía op_203 (fijo E14) y op_204 (copia de op_200). Para
        # el resto de muebles, AS400 sigue calculándolas — no se envían (sin cambios).
        if code in codigos_tapeta:
            sg_203 = (forzadas_yaml.get("op_203") or {}).get("sg", "E14")
            _sg("op_203", sg_203)
            if sg_200:
                _sg("op_204", sg_200)

    # ── op_206 — Sistema de cierre FSP (F en muebles batientes) ─────────────
    if not es_abierto and code in (op_mueble.get("op_206") or []):
        sg_206 = (forzadas_yaml.get("op_206") or {}).get("sg", "FSP")
        _sg("op_206", sg_206)

    # ── op_121 — Sin mecanizado para tirador ─────────────────────────────────
    # Forzado por reglas (Plantea o Curve/Line en monoporte) → origen automático
    # Seleccionado por el usuario → origen usuario
    if _es_true(fila.get("Sin mecanizado")):
        ui_to_skp       = indices.get("op_300_ui_to_skp") or {}
        tirador_skp     = ui_to_skp.get(tirador_ui, "")
        r121_c          = reglas_c.get("op_121") or {}
        r121_b          = reglas_b.get("op_121") or {}
        forzado_siempre = [str(t) for t in (r121_c.get("tirador_spf_siempre") or [])]
        forzado_mono_t  = [str(t) for t in (r121_c.get("tiradores_spf_forzado_monoporte") or [])]
        forzado_mono_m  = (r121_b.get("forzado_en_monoporte") or {}).get("muebles") or []

        es_forzado = (
            tirador_skp in forzado_siempre
            or (tirador_skp in forzado_mono_t and code in forzado_mono_m)
        )
        _sg(
            "op_121", "SPF",
            origen="automatico" if es_forzado else "usuario",
            etiqueta="Sin mecanizado para tirador",
        )
        if es_forzado:
            avisos.append("Este mueble no lleva mecanización para tirador.")

    # ── op_207 — Equipamiento fregadero / cubos de basura / despensa ─────────
    datos_207 = op_mueble.get("op_207") or {}
    if code in (datos_207.get("P42") or []):
        # Forzado: fregadero equipado (BETOQ908057)
        _sg("op_207", "P42", origen="automatico", etiqueta="Equipamiento fregadero")
    else:
        val_207 = (fila.get("Cubos de basura") or "").strip()
        if val_207 in ("GM1", "GM2"):
            # Despensa AGM: el usuario seleccionó tipo de almacenamiento
            if code in (datos_207.get(val_207) or []):
                _sg("op_207", val_207, origen="usuario", etiqueta="Tipo de almacenamiento")
        elif _es_true(val_207):
            if code in (datos_207.get("P60") or []):
                _sg("op_207", "P60", origen="usuario", etiqueta="Cubos de basura")
            elif code in (datos_207.get("P90") or []):
                _sg("op_207", "P90", origen="usuario", etiqueta="Cubos de basura")

    # ── op_208 — Equipamiento esquinero Lemans (F) ───────────────────────────
    if code in (op_mueble.get("op_208") or []):
        sg_208 = (forzadas_yaml.get("op_208") or {}).get("sg", "CPA")
        _sg("op_208", sg_208, origen="automatico", etiqueta="Equipamiento esquinero")

    # ── op_209 — EXCLUIDA: SG la calcula en AS400 ───────────────────────────

    # ── op_217 — Sistema de apertura sin tirador físico ──────────────────────
    if not es_abierto:
        datos_217 = op_mueble.get("op_217") or {}
        if tirador_ui == "Touch Latch" and code in (datos_217.get("solo_TL1") or []):
            _sg("op_217", "TL1")
        elif tirador_ui == "Prise de main" and code in (datos_217.get("solo_PS1") or []):
            _sg("op_217", "PS1")

    # ── op_220 — Recorte para perfil LED ─────────────────────────────────────
    if code in (op_mueble.get("op_220") or []) and _es_true(fila.get("Recorte LED")):
        _sg("op_220", "PRB", origen="usuario", etiqueta="Recorte para perfil LED")

    # ── op_222 — Sensor para mando LED ───────────────────────────────────────
    sensor = (fila.get("Sensor para mando LED") or "").strip().lower()
    if code in (op_mueble.get("op_222") or []):
        if sensor == "derecha":
            _sg("op_222", "DTD", origen="usuario", etiqueta="Sensor mando LED")
        elif sensor == "izquierda":
            _sg("op_222", "DTG", origen="usuario", etiqueta="Sensor mando LED")

    # ── op_223 — Cajón interior ───────────────────────────────────────────────
    if code in (op_mueble.get("op_223") or []) and _es_true(fila.get("Cajón interior")):
        _sg("op_223", "TLH", origen="usuario", etiqueta="Cajón interior")

    # ── op_227 — Mueble de caldera ────────────────────────────────────────────
    if code in (op_mueble.get("op_227") or []) and _es_true(fila.get("Mueble de caldera")):
        _sg("op_227", "CCH", origen="usuario", etiqueta="Mueble de caldera")

    # ── op_231 — Reducción de ancho ───────────────────────────────────────────
    if code in (op_mueble.get("op_231") or []) and _es_true(fila.get("Reducción de ancho")):
        ancho_raw = (fila.get("Ancho reducido") or "").strip()
        valor_mm  = ancho_raw.replace("mm", "").replace(" ", "").strip()
        _sg("op_231", "RL3", param1=valor_mm, origen="usuario", etiqueta="Reducción de ancho")

    # ── op_300 — Tipo de tirador (todos: O, excepto muebles abiertos) ───────────
    if not es_abierto:
        sg_300_base = indices.get("op_300", {}).get(tirador_ui, "")
        if sg_300_base:
            if tirador_ui in ("Curve", "Line"):
                # Sufijo H (batiente) o C (coulissant/extraíble/banco)
                sufijo  = "H" if _es_batiente(code, op_mueble) else "C"
                prefijo = "Q2" if tirador_ui == "Curve" else "Q3"
                sg_300  = f"{prefijo}{sufijo}"
            else:
                sg_300 = sg_300_base
            _sg("op_300", sg_300)

        # ── op_301 — Color del tirador ────────────────────────────────────────────
        excl_301 = (op_mueble.get("op_301") or {}).get("excepciones") or []
        if code not in excl_301 and tirador_ui not in _TIRADORES_SIN_COLOR and tirador_ui:
            color = (fila.get("Color tirador") or "").strip()
            if tirador_ui in _TIRADORES_INTEGRADOS:
                # Round/Square: color de la trasera (o del frente si Trasera=Laca)
                # ya resuelto por Módulo B en el campo "Color tirador"
                sg_301 = indices.get("op_301_integrado", {}).get(color, "")
            else:
                # Curve/Line/Plantea: color del tirador de superficie
                sg_301 = indices.get("op_301_superficie", {}).get(color, "")
            if sg_301:
                _sg("op_301", sg_301)

    # ── op_402 — Rodapié / altura de patas ────────────────────────────────────
    excl_402 = (op_mueble.get("op_402") or {}).get("excepciones") or []
    if code not in excl_402:
        rodapie = (fila.get("Rodapié") or "").strip()
        sg_402  = indices.get("op_402", {}).get(rodapie, "")
        if sg_402:
            _sg("op_402", sg_402)

    # ── op_700 — Mueble sin encolar ────────────────────────────────────────────
    datos_700   = op_mueble.get("op_700") or {}
    forzados700 = datos_700.get("forzadas")  or []
    no_aplica700 = datos_700.get("no_aplica") or []
    if code in forzados700:
        # HH*: DEM siempre obligatorio (muebles en escuadra, sin encolar para montaje)
        _sg("op_700", "DEM", origen="automatico", etiqueta="Mueble sin encolar")
        avisos.append("Este mueble siempre se entrega sin encolar para facilitar el montaje.")
    elif code not in no_aplica700 and _es_true(fila.get("Sin encolar")):
        # Resto: opcional, solo si el usuario lo marcó
        _sg("op_700", "DEM", origen="usuario", etiqueta="Mueble sin encolar")

    return opciones_sg, opc_adicionales, codigos, avisos


# =============================================================================
# Función principal
# =============================================================================

def calcular_opciones(entrada: list[dict]) -> list[dict]:
    """
    Módulo C — función principal.

    Input:  list[dict] — 23 columnas (CLAUDE.md §9), valores en lenguaje UI.
    Output: list[dict] — pedido extendido para modulo_b.paso_2():
              · todos los campos originales de entrada
              · opciones_adicionales: list[dict] {etiqueta, valor, origen}
              · codigos_sg:           dict {op_id → código_sg}
              · p_item:               dict en formato p_items de la API SG
              · avisos_c:             list[str] de códigos AV0x
    """
    if not entrada:
        return []

    catalogo, mapeos, op_mueble, reglas = _cargar_datos()
    indices    = _build_indices(mapeos)
    nombres_fr = mapeos.get("nombres") or {}

    # Códigos SG de rodapié (SOC36010, SOC18010, SOC3607, SOC1807) — para p_quantity
    _codigos_rodapie_sg: set[str] = set((op_mueble.get("rodapiés") or {}).get("codigos_sg") or [])

    resultado: list[dict] = []

    for i, fila in enumerate(entrada):
        code = (fila.get("Código mueble") or "").strip()

        # ── Designación en francés ────────────────────────────────────────────
        cat_entry = catalogo.get(code) or {}
        des_es    = (cat_entry.get("designaciones") or {}).get("es", "")
        # Buscar en la tabla de nombres del YAML; si no existe, usar el campo "fr" del catálogo
        label_fr  = nombres_fr.get(des_es, "") or (cat_entry.get("designaciones") or {}).get("fr", "")

        # ── p_hinge ───────────────────────────────────────────────────────────
        _codigos_abierto: set[str] = set(
            ((op_mueble.get("op_410") or {}).get("codigos")) or []
        )
        _codigos_joue_hinge: set[str] = set((op_mueble.get("joues") or {}).get("codigos") or [])
        if code in _codigos_abierto or code in _codigos_joue_hinge:
            p_hinge: str | None = None  # muebles abiertos y joues: sin p_hinge en el JSON
        else:
            cat_hinge   = _p_hinge_cat(code, op_mueble)
            apertura_ui = (fila.get("Apertura") or "").strip()
            if cat_hinge == "lee_csv":
                p_hinge = indices.get("apertura", {}).get(apertura_ui)
            elif cat_hinge == "nulo":
                p_hinge = None  # 2 puertas horizontales — p_hinge no aplica, se omite del JSON
            elif cat_hinge == "coulissant":
                p_hinge = "C"
            elif cat_hinge == "lift":
                p_hinge = "L"   # apertura lift — confirmado con SG
            else:
                p_hinge = None

        # ── p_fastening ───────────────────────────────────────────────────────
        # Tapetas, rodapiés y joues: no se envía. Se omite del JSON vía _sin_nulos.
        _codigos_tapeta_fast: set[str] = set((op_mueble.get("tapetas") or {}).get("codigos") or [])
        _codigos_rodapie_fast: set[str] = set((op_mueble.get("rodapiés") or {}).get("codigos") or [])
        _codigos_joue_fast: set[str] = set((op_mueble.get("joues") or {}).get("codigos") or [])
        if code in _codigos_tapeta_fast or code in _codigos_rodapie_fast or code in _codigos_joue_fast:
            p_fastening = None
        else:
            _ps_segun_rodapie: set[str] = set(
                ((op_mueble.get("op_402") or {}).get("p_s_segun_rodapie")) or []
            )
            if code in _ps_segun_rodapie:
                # EOVV37, EOAVV37: suspendido si rodapié vacío, posado si "70 mm" o "100 mm"
                _rodapie_val = (fila.get("Rodapié") or "").strip()
                p_fastening = "P" if _rodapie_val else "S"
            else:
                p_fastening = "S" if _es_suspendido(code, op_mueble) else "P"

        # Si viene "Posición" del CSV, sobreescribe p_fastening.
        # Tapetas: B o H. Muebles abiertos EO*: P o S (reemplaza el cálculo automático).
        _posicion_csv = (fila.get("Posición") or "").strip().upper()
        if _posicion_csv in ("B", "H", "P", "S"):
            p_fastening = _posicion_csv

        # ── Opciones ──────────────────────────────────────────────────────────
        opciones_sg, opc_adic, codigos, avisos = _calcular_opciones_mueble(
            fila, mapeos, op_mueble, reglas, indices
        )

        # ── Aviso informativo: muebles que siempre se envían desmontados ──────
        _av_d        = op_mueble.get("avisos_desmontado") or {}
        _av_codigos  = set(_av_d.get("codigos")  or [])
        _av_prefijos = tuple(_av_d.get("prefijos") or [])
        if code in _av_codigos or (_av_prefijos and code.startswith(_av_prefijos)):
            avisos.append("Este mueble siempre se envía desmontado al cliente.")

        # ── p_built_in_detail ─────────────────────────────────────────────────
        # Soporta hasta 2 electros (AFSMO/AFSMOBT): electro 1 = inferior, electro 2 = superior.
        # Caso A (referencia conocida): p_appliance_reference.
        # Caso B (sin referencia): p_built_in_height (string, en mm).
        _electros = [
            (
                (fila.get("Marca electro")        or "").strip(),
                (fila.get("Referencia electro")   or "").strip(),
                (fila.get("Alto electro")         or "").strip(),
            ),
            (
                (fila.get("Marca electro 2")      or "").strip(),
                (fila.get("Referencia electro 2") or "").strip(),
                (fila.get("Alto electro 2")       or "").strip(),
            ),
        ]
        p_built_in_entries: list[dict] = []
        for m, r, al in _electros:
            if r:
                # Caso A: referencia conocida — Marca + Referencia
                # p_manufacturer_code: 3 primeras letras de la marca, en mayúsculas
                # (ej. "Siemens" → "SIE"). La marca completa se sigue mostrando tal
                # cual en Configuración/PDF — solo el JSON usa el código abreviado.
                p_built_in_entries.append({
                    "p_manufacturer_code":   m.strip()[:3].upper(),
                    "p_appliance_reference": r,
                })
            elif al:
                # Caso B: sin referencia — solo altura en mm (sin marca)
                p_built_in_entries.append({
                    "p_built_in_height": str(al),
                })
        p_built_in: list[dict] | None = p_built_in_entries or None

        # ── Item JSON ─────────────────────────────────────────────────────────
        # Los campos estáticos (p_quantity, p_width/height/depth, p_delivery_date…)
        # vienen de data/p_item_schema.yaml; los calculados se construyen aquí.
        # El orden de claves sigue exactamente el contrato de Schmidt Groupe.
        _d = _p_item_defaults()

        # Para piezas con dimensiones variables, p_width/p_height se toman del CSV.
        # Piezas con alto_variable: HR*, HLVV57, HPT60V57, FF12V, J19VV.
        # Piezas con ancho_variable: HLVV57, J19VV.
        # SG necesita las dimensiones reales para fabricar; las fijas las conoce por código.
        _alto_final  = (fila.get("Alto final tapeta") or "").strip()
        _ancho_csv   = (fila.get("Ancho CSV") or "").replace("mm", "").strip()
        _alto_csv    = (fila.get("Alto CSV")  or "").replace("mm", "").strip()
        _cat_e_dim   = catalogo.get(code) or {}
        _tiene_alto_var  = _cat_e_dim.get("alto_variable")  is not None or _cat_e_dim.get("alto_variable_por_gama")  is not None
        _tiene_ancho_var = _cat_e_dim.get("ancho_variable") is not None or _cat_e_dim.get("ancho_variable_por_gama") is not None

        # Dimensiones: siempre presentes. Variable → valor del CSV. Fijo → 0.
        _codigos_enc_dim: set[str] = (
            set((op_mueble.get("encimeras_panel") or {}).get("codigos") or []) |
            set((op_mueble.get("encimeras_pdt")   or {}).get("codigos") or [])
        )
        if code == "FF12V":
            _p_width  = 0
            _p_height = int(_alto_final) if _alto_final.isdigit() else 0
            _p_depth  = 0
        elif code in _codigos_enc_dim:
            _p_height = int(_ancho_csv) if _ancho_csv.isdigit() else 0
            _p_width  = int(_alto_csv)  if _alto_csv.isdigit()  else 0
            _p_depth  = 0
        else:
            _p_width  = int(_ancho_csv) if _tiene_ancho_var and _ancho_csv.isdigit() else 0
            _p_height = int(_alto_csv)  if _tiene_alto_var  and _alto_csv.isdigit()  else 0
            _p_depth  = 0

        p_item: dict = {
            "p_ord_cat_code":          str(i + 1),
            "p_item_code":             code,
            "p_item_label":            label_fr,
            "p_item_origin_id":        (fila.get("Summary") or "").strip() or None,
            "p_quantity":              int((fila.get("Cantidad") or "").strip() or 0) if code in _codigos_rodapie_sg and (fila.get("Cantidad") or "").strip() else _d.get("p_quantity", 1),
            "p_hinge":                 p_hinge,
            "p_fastening":             p_fastening,
            "p_width":                 _p_width,
            "p_height":                _p_height,
            "p_depth":                 _p_depth,
            "p_delivery_date":         _d.get("p_delivery_date"),
            "p_variant_options":       opciones_sg,
            **( {"p_built_in_detail": p_built_in} if p_built_in else {} ),
        }

        resultado.append({
            **fila,                              # preserva las 23 columnas de entrada
            "opciones_adicionales": opc_adic,
            "codigos_sg":           codigos,
            "p_item":               p_item,
            "avisos_c":             avisos,
        })

    return resultado


# =============================================================================
# Export: JSON de pedido (p_items)
# =============================================================================

def _sin_nulos(obj):
    """Elimina recursivamente las claves cuyo valor es None (null en JSON).
    Valores falsy distintos de None (0, "", False, []) se conservan."""
    if isinstance(obj, dict):
        return {k: _sin_nulos(v) for k, v in obj.items() if v is not None}
    if isinstance(obj, list):
        return [_sin_nulos(i) for i in obj]
    return obj


def generar_json_pedido(resultado: list[dict]) -> list[dict]:
    """
    Extrae la lista p_items del resultado de calcular_opciones.
    Solo esta parte se envía a la API de Schmidt Groupe.
    Los campos con valor null se eliminan del output.
    """
    return [_sin_nulos(r["p_item"]) for r in resultado if "p_item" in r]


# =============================================================================
# Export: Excel de revisión
# =============================================================================

# Definición de columnas del Excel.
# Cada entrada: (header, tipo, key_entrada, key_salida)
#   tipo:        "input" | "output" | "aviso"
#   key_entrada: campo del dict original (None si es columna solo output)
#   key_salida:  clave en codigos_sg, o clave especial "_xxx" (None si input puro)
_EXCEL_COLS: list[tuple[str, str, str | None, str | None]] = [
    ("Código mueble",       "input",  "Código mueble",         None),
    ("p_item_code",         "output", None,                    "_p_item_code"),
    ("Descripción",         "input",  "Descripción",           None),
    ("p_item_label (FR)",   "output", None,                    "_p_item_label"),
    ("Apertura CSV",        "input",  "Apertura",              None),
    ("p_hinge",             "output", None,                    "_p_hinge"),
    ("p_fastening",         "output", None,                    "_p_fastening"),
    ("Gama del frente",     "input",  "Gama del frente",       None),
    ("op_100",              "output", None,                    "op_100"),
    ("Acabado del frente",  "input",  "Acabado del frente",    None),
    ("op_101",              "output", None,                    "op_101"),
    ("op_621",              "output", None,                    "op_621"),
    ("Color interior",      "input",  "Color interior",        None),
    ("op_200",              "output", None,                    "op_200"),
    ("op_206",              "output", None,                    "op_206"),
    ("Tirador CSV",         "input",  "Tirador",               None),
    ("op_217",              "output", None,                    "op_217"),
    ("op_300",              "output", None,                    "op_300"),
    ("Color tirador CSV",   "input",  "Color tirador",         None),
    ("op_301",              "output", None,                    "op_301"),
    ("Sin mecanizado CSV",  "input",  "Sin mecanizado",        None),
    ("op_121",              "output", None,                    "op_121"),
    ("Reducción CSV",       "input",  "Reducción de ancho",    None),
    ("op_231 art.",         "output", None,                    "op_231"),
    ("Ancho reducido CSV",  "input",  "Ancho reducido",        None),
    ("op_231 p_param1",     "output", None,                    "op_231_param1"),
    ("Recorte LED CSV",     "input",  "Recorte LED",           None),
    ("op_220",              "output", None,                    "op_220"),
    ("Sensor CSV",          "input",  "Sensor para mando LED", None),
    ("op_222",              "output", None,                    "op_222"),
    ("Cajón interior CSV",  "input",  "Cajón interior",        None),
    ("op_223",              "output", None,                    "op_223"),
    ("Mueble caldera CSV",  "input",  "Mueble de caldera",     None),
    ("op_227",              "output", None,                    "op_227"),
    ("Cubos de basura CSV", "input",  "Cubos de basura",       None),
    ("op_207",              "output", None,                    "op_207"),
    ("Sin encolar CSV",     "input",  "Sin encolar",           None),
    ("op_700",              "output", None,                    "op_700"),
    ("Rodapié",             "input",  "Rodapié",               None),
    ("op_402",              "output", None,                    "op_402"),
    ("Marca electro",       "input",  "Marca electro",         None),
    ("p_built_in_manuf.",   "output", None,                    "_manuf"),
    ("Referencia electro",  "input",  "Referencia electro",    None),
    ("p_built_in_ref.",     "output", None,                    "_ref"),
    ("Avisos",              "aviso",  None,                    "_avisos"),
]


def _valor_excel(res: dict, key_entrada: str | None, key_salida: str | None) -> str:
    """Extrae el valor de celda para el Excel a partir de un item del resultado."""
    # res contiene todos los campos de entrada + los campos calculados
    p_item  = res.get("p_item") or {}
    codigos = res.get("codigos_sg") or {}
    bid     = (p_item.get("p_built_in_detail") or [{}])[0]
    avisos  = res.get("avisos_c") or []

    if key_entrada is not None:
        return str(res.get(key_entrada) or "")

    especiales: dict[str, str] = {
        "_p_item_code":  p_item.get("p_item_code", ""),
        "_p_item_label": p_item.get("p_item_label", ""),
        "_p_hinge":      "" if p_item.get("p_hinge") is None else str(p_item["p_hinge"]),
        "_p_fastening":  p_item.get("p_fastening", ""),
        "_manuf":        bid.get("p_manufacturer_code", "") or "",
        "_ref":          bid.get("p_appliance_reference", "") or "",
        "_avisos":       " | ".join(avisos),
    }
    if key_salida in especiales:
        return str(especiales[key_salida])

    return str(codigos.get(key_salida, "") or "")


def generar_excel_revision(resultado: list[dict]) -> bytes:
    """
    Genera el Excel de revisión a partir del resultado de calcular_opciones.

    Formato:
        Fila 1: título (fondo negro, texto blanco)
        Fila 2: cabeceras (verde=input, azul=output, naranja=avisos)
        Fila 3+: datos con los mismos colores por columna
        Paneles congelados en A3.

    Requiere openpyxl (añadido a requirements.txt).
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ImportError(
            "openpyxl no está instalado. Añádelo a requirements.txt."
        ) from exc

    # Paleta
    def _fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color)

    FILL_HDR_INPUT  = _fill("1A5C2A")
    FILL_HDR_OUTPUT = _fill("1F3864")
    FILL_HDR_AVISO  = _fill("FF8C00")
    FILL_HDR_TIT    = _fill("1F1F1F")
    FILL_DAT_INPUT  = _fill("EBF5E9")
    FILL_DAT_OUTPUT = _fill("E8F0FE")
    FILL_DAT_AVISO  = _fill("FFF3E0")

    FONT_TIT  = Font(bold=True, color="FFFFFF", size=11)
    FONT_HDR  = Font(bold=True, color="FFFFFF", size=10)
    FONT_DAT  = Font(size=10)

    BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    ALIGN_CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ALIGN_LFT = Alignment(horizontal="left",   vertical="center", wrap_text=False)

    # Mapas de fills por tipo
    _hdr_fill = {"input": FILL_HDR_INPUT, "output": FILL_HDR_OUTPUT, "aviso": FILL_HDR_AVISO}
    _dat_fill = {"input": FILL_DAT_INPUT, "output": FILL_DAT_OUTPUT, "aviso": FILL_DAT_AVISO}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Revisión"

    n_cols = len(_EXCEL_COLS)

    # Fila 1 — título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    cell = ws.cell(row=1, column=1, value="CUBRO × Schmidt Groupe — Revisión de pedido")
    cell.fill      = FILL_HDR_TIT
    cell.font      = FONT_TIT
    cell.alignment = ALIGN_CTR
    ws.row_dimensions[1].height = 24

    # Fila 2 — cabeceras
    for ci, (header, tipo, _, _2) in enumerate(_EXCEL_COLS, start=1):
        cell = ws.cell(row=2, column=ci, value=header)
        cell.fill      = _hdr_fill[tipo]
        cell.font      = FONT_HDR
        cell.alignment = ALIGN_CTR
        cell.border    = BORDER
    ws.row_dimensions[2].height = 32

    # Filas de datos
    for ri, res in enumerate(resultado, start=3):
        for ci, (_, tipo, key_ent, key_sal) in enumerate(_EXCEL_COLS, start=1):
            valor = _valor_excel(res, key_ent, key_sal)
            cell  = ws.cell(row=ri, column=ci, value=valor)
            cell.fill      = _dat_fill[tipo]
            cell.font      = FONT_DAT
            cell.alignment = ALIGN_LFT
            cell.border    = BORDER
        ws.row_dimensions[ri].height = 16

    # Congelar desde fila 3 (primera fila de datos)
    ws.freeze_panes = "A3"

    # Anchos de columna
    _ANCHOS = {
        "Código mueble": 18, "p_item_code": 18,
        "Descripción": 32, "p_item_label (FR)": 40,
        "Avisos": 55,
    }
    default_w = 14
    for ci, (header, _, _, _2) in enumerate(_EXCEL_COLS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = _ANCHOS.get(header, default_w)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
